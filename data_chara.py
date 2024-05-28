import openpyxl
import shutil

esl_lesson_plans = [
   {
       "section": "Daily Routines",
       "learning_target": "describing daily activities",
       "actual_content": "I wake up at 7 a.m. every morning."
   },
   {
       "section": "Family and Friends",
       "learning_target": "describing relationships",
       "actual_content": "My best friend is always there for me."
   },
   {
       "section": "Hobbies and Interests",
       "learning_target": "discussing hobbies",
       "actual_content": "I enjoy painting landscapes in my free time."
   },
   {
       "section": "Travel Experiences",
       "learning_target": "describing travel destinations",
       "actual_content": "The beaches in Hawaii were breathtaking."
   },
   {
       "section": "Favorite Foods",
       "learning_target": "discussing food preferences",
       "actual_content": "I can't resist a slice of homemade apple pie."
   },
   {
       "section": "Job and Career",
       "learning_target": "discussing career goals",
       "actual_content": "My dream job is to become a software engineer."
   },
   {
       "section": "Health and Fitness",
       "learning_target": "describing exercise routines",
       "actual_content": "I go for a jog every evening after work."
   },
   {
       "section": "Shopping Experiences",
       "learning_target": "describing shopping trips",
       "actual_content": "I found a great deal on a new laptop."
   },
   {
       "section": "Environmental Issues",
       "learning_target": "discussing environmental concerns",
       "actual_content": "We should all do our part to reduce plastic waste."
   },
   {
       "section": "Cultural Traditions",
       "learning_target": "describing cultural celebrations",
       "actual_content": "Chinese New Year is celebrated with fireworks and parades."
   },
   {
       "section": "Local Attractions",
       "learning_target": "describing places of interest",
       "actual_content": "The Golden Gate Bridge is a must-see landmark."
   },
   {
       "section": "Personal Goals",
       "learning_target": "discussing future plans",
       "actual_content": "My goal is to learn a new language this year."
   },
   {
       "section": "School and Education",
       "learning_target": "discussing academic experiences",
       "actual_content": "I struggled with calculus in my first year of college."
   },
   {
       "section": "Favorite Movies and Books",
       "learning_target": "describing entertainment preferences",
       "actual_content": "I recently read a captivating novel by a new author."
   },
   {
       "section": "Sports and Recreation",
       "learning_target": "discussing physical activities",
       "actual_content": "I play soccer with my friends every weekend."
   },
   {
       "section": "Technology and Gadgets",
       "learning_target": "describing technology experiences",
       "actual_content": "The new smartphone has an impressive camera."
   },
   {
       "section": "Home and Living",
       "learning_target": "describing living situations",
       "actual_content": "I live in a cozy apartment in the city center."
   },
   {
       "section": "Fashion and Style",
       "learning_target": "discussing fashion preferences",
       "actual_content": "I prefer casual and comfortable clothing."
   },
   {
       "section": "Music and Dance",
       "learning_target": "describing musical interests",
       "actual_content": "I enjoy listening to classical music while studying."
   },
   {
       "section": "Current Events",
       "learning_target": "discussing news and events",
       "actual_content": "The recent election results were surprising."
   },
   {
       "section": "Art and Culture",
       "learning_target": "describing artistic experiences",
       "actual_content": "I visited an impressive art exhibition last weekend."
   },
   {
       "section": "Travel Planning",
       "learning_target": "discussing travel preparations",
       "actual_content": "I'm researching affordable hotels for my upcoming trip."
   },
   {
       "section": "Health and Nutrition",
       "learning_target": "discussing dietary habits",
       "actual_content": "I try to incorporate more vegetables into my diet."
   },
   {
       "section": "Pet Care",
       "learning_target": "describing pet experiences",
       "actual_content": "My dog loves playing fetch in the park."
   },
   {
       "section": "Outdoor Adventures",
       "learning_target": "discussing outdoor activities",
       "actual_content": "Hiking in the mountains is a great way to stay active."
   }
]

print(len(esl_lesson_plans))


def insert_lesson_plans(filename, lesson_plans, start_row=9):
  """
  Inserts lesson plans into the original Excel document.

  Args:
    filename: The name of the original Excel file.
    lesson_plans: A list of dictionaries containing lesson plans.
    start_row: The starting row (inclusive) to insert data (default: 9).
  """
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    print(f"File not found: {filename}")
    return

  worksheet = workbook.active

  # Insert lesson plans starting from start_row
  row_index = start_row
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=4).value = "50 mins"
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  # Save the changes to the original document
  workbook.save(filename)
  print(f"Lesson plans inserted. Original file modified: {filename}")


# Example usage
filename = "[TLI]_Blizzard_Learning_program_Chara.xlsx"

# Remember to replace this with your actual backup directory path
#backup_directory = "/home/sweeneyphilip11/TLI_Schedule/backup_directory"
backup_directory = "/home/thinkpad/Documents/Python_Excel_Automation/Excel_Automation_AI_Practice/backup_directory"


# Create a backup of the original file (consider using libraries like shutil for robust copying)
shutil.copy(filename, f"{backup_directory}/{filename}")  # Example using shutil.copy

insert_lesson_plans(filename, esl_lesson_plans)


