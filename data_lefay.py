import openpyxl
import shutil

esl_conversation_topics = [
   {
       "section": "Family Life",
       "learning_target": "Discussing family traditions and values",
       "actual_content": "In my culture, respecting elders is very important."
   },
   {
       "section": "Hobbies and Interests",
       "learning_target": "Sharing personal hobbies and interests",
       "actual_content": "I've been playing guitar since I was a teenager."
   },
   {
       "section": "Travel Experiences",
       "learning_target": "Exchanging travel stories and recommendations",
       "actual_content": "My favorite place I've visited is Tokyo."
   },
   {
       "section": "Cultural Diversity",
       "learning_target": "Exploring different cultures and traditions",
       "actual_content": "I find the diversity of customs fascinating."
   },
   {
       "section": "Environmental Concerns",
       "learning_target": "Discussing environmental issues and solutions",
       "actual_content": "Recycling is an easy way to reduce waste."
   },
   {
       "section": "Career and Job",
       "learning_target": "Sharing career goals and job experiences",
       "actual_content": "I'm considering a career change to teaching."
   },
   {
       "section": "Personal Goals",
       "learning_target": "Discussing personal goals and aspirations",
       "actual_content": "One of my goals is to learn a third language."
   },
   {
       "section": "Health and Fitness",
       "learning_target": "Exchanging health and fitness tips",
       "actual_content": "I've found that yoga helps me manage stress."
   },
   {
       "section": "Food and Cooking",
       "learning_target": "Sharing favorite recipes and cooking experiences",
       "actual_content": "My grandmother's secret ingredient is love."
   },
   {
       "section": "Entertainment",
       "learning_target": "Discussing movies, books, and music preferences",
       "actual_content": "I'm a fan of classic rock music from the 70s."
   },
   {
       "section": "Technology and Innovation",
       "learning_target": "Exploring new technologies and their impact",
       "actual_content": "Virtual reality gaming is the future of entertainment."
   },
   {
       "section": "Sports and Recreation",
       "learning_target": "Sharing experiences with sports and outdoor activities",
       "actual_content": "I enjoy hiking on weekends to unwind."
   },
   {
       "section": "Current Events",
       "learning_target": "Discussing news and current affairs",
       "actual_content": "The recent political scandal has been widely debated."
   },
   {
       "section": "Art and Culture",
       "learning_target": "Appreciating different forms of art and cultural expression",
       "actual_content": "I find abstract paintings intriguing and thought-provoking."
   },
   {
       "section": "Education and Learning",
       "learning_target": "Exchanging perspectives on education systems",
       "actual_content": "I believe hands-on learning is more effective."
   },
   {
       "section": "Travel and Adventure",
       "learning_target": "Sharing stories of adventure and exploration",
       "actual_content": "I once went skydiving and it was exhilarating."
   },
   {
       "section": "Fashion and Style",
       "learning_target": "Discussing fashion trends and personal style",
       "actual_content": "I prefer minimalist and timeless fashion choices."
   },
   {
       "section": "Home and Living",
       "learning_target": "Exchanging tips on home organization and decor",
       "actual_content": "I love incorporating plants into my living space."
   },
   {
       "section": "Relationships and Dating",
       "learning_target": "Sharing perspectives on dating and relationships",
       "actual_content": "Communication is key to a healthy relationship."
   },
   {
       "section": "Parenting and Child-rearing",
       "learning_target": "Discussing parenting challenges and experiences",
       "actual_content": "Setting a good example is crucial for children."
   },
   {
       "section": "Mental Health and Well-being",
       "learning_target": "Exploring strategies for mental wellness",
       "actual_content": "Practicing mindfulness has been beneficial for me."
   },
   {
       "section": "Business and Entrepreneurship",
       "learning_target": "Sharing business ideas and experiences",
       "actual_content": "Starting a small business requires passion and perseverance."
   },
   {
       "section": "Language Learning",
       "learning_target": "Exchanging language learning tips and experiences",
       "actual_content": "Immersion is the best way to learn a new language."
   },
   {
       "section": "Social Issues",
       "learning_target": "Discussing social challenges and solutions",
       "actual_content": "Poverty and inequality remain major global issues."
   },
   {
       "section": "Science and Technology",
       "learning_target": "Exploring scientific discoveries and technological advancements",
       "actual_content": "Renewable energy sources are crucial for a sustainable future."
   }
]

print(len(esl_conversation_topics))

def insert_lesson_plans(filename, lesson_plans, start_row=9):
  """
  Inserts lesson plans into the original Excel document.

  Args:
    filename: The name of the original Excel file.
    lesson_plans: A list of dictionaries containing lesson plans.
    start_row: The starting row (inclusive) to insert data (default: 9).
  """
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    print(f"File not found: {filename}")
    return

  worksheet = workbook.active

  # Insert lesson plans starting from start_row
  row_index = start_row
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=4).value = "50 mins"
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  # Save the changes to the original document
  workbook.save(filename)
  print(f"Lesson plans inserted. Original file modified: {filename}")


# Example usage
filename = "[TLI]_Blizzard_Learning_program_Lefay.xlsx"

# Remember to replace this with your actual backup directory path
#backup_directory = "/home/sweeneyphilip11/TLI_Schedule/backup_directory"
backup_directory = "/home/thinkpad/Documents/Python_Excel_Automation/Excel_Automation_AI_Practice/backup_directory"


# Create a backup of the original file (consider using libraries like shutil for robust copying)
shutil.copy(filename, f"{backup_directory}/{filename}")  # Example using shutil.copy

insert_lesson_plans(filename, esl_conversation_topics)


def insert_lesson_plans(filename, lesson_plans, start_row=9):
  """
  Inserts lesson plans into the original Excel document.

  Args:
    filename: The name of the original Excel file.
    lesson_plans: A list of dictionaries containing lesson plans.
    start_row: The starting row (inclusive) to insert data (default: 9).
  """
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    print(f"File not found: {filename}")
    return

  worksheet = workbook.active

  # Insert lesson plans starting from start_row
  row_index = start_row
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=4).value = "50 mins"
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  # Save the changes to the original document
  workbook.save(filename)
  print(f"Lesson plans inserted. Original file modified: {filename}")


# Example usage
filename = "[TLI]_Blizzard_Learning_program_Lefay.xlsx"

# Remember to replace this with your actual backup directory path
#backup_directory = "/home/sweeneyphilip11/TLI_Schedule/backup_directory"
backup_directory = "/home/thinkpad/Documents/Python_Excel_Automation/Excel_Automation_AI_Practice/backup_directory"


# Create a backup of the original file (consider using libraries like shutil for robust copying)
shutil.copy(filename, f"{backup_directory}/{filename}")  # Example using shutil.copy

insert_lesson_plans(filename, esl_conversation_topics)
