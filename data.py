import openpyxl
import shutil

esl_lesson_plans = [
  { "section": "Vocabulary", "learning_target": "describing places", "actual_content": "big/small, clean/dirty" },
  { "section": "Grammar", "learning_target": "present continuous", "actual_content": "am/is/are doing" },
  { "section": "Reading Comprehension", "learning_target": "main idea", "actual_content": "who/what/where" },
  { "section": "Listening Comprehension", "learning_target": "numbers", "actual_content": "one/two/three" },
  { "section": "Speaking", "learning_target": "greetings", "actual_content": "Hello/Goodbye" },
  { "section": "Writing", "learning_target": "simple sentences", "actual_content": "Subject + verb" },

  # Alternate sections

  { "section": "Vocabulary", "learning_target": "describing people", "actual_content": "tall/short, happy/sad" },
  { "section": "Grammar", "learning_target": "present simple tense", "actual_content": "he/she/it + verb" },
  { "section": "Reading Comprehension", "learning_target": "sequencing events", "actual_content": "first/then/next" },
  { "section": "Listening Comprehension", "learning_target": "identifying simple questions", "actual_content": "What/Where/When" },
  { "section": "Speaking", "learning_target": "introducing yourself", "actual_content": "Hi, I'm..." },
  { "section": "Writing", "learning_target": "using punctuation (period, comma)", "actual_content": ". , ?" },

  # Continue alternating sections

  { "section": "Vocabulary", "learning_target": "food and drinks", "actual_content": "apple/juice, water/coffee" },
  { "section": "Grammar", "learning_target": "yes/no questions", "actual_content": "Do/Does + verb?" },
  { "section": "Reading Comprehension", "learning_target": "identifying characters", "actual_content": "who/what" },
  { "section": "Listening Comprehension", "learning_target": "following instructions", "actual_content": "Listen and do" },
  { "section": "Speaking", "learning_target": "asking and answering questions", "actual_content": "What's your name?/My name is..." },
  { "section": "Writing", "learning_target": "articles (a/an/the)", "actual_content": "a book, the cat" },

  # Last two items

  { "section": "Vocabulary", "learning_target": "colors", "actual_content": "red/blue, yellow/green" },
  # Three additional items
  { "section": "Grammar", "learning_target": "past tense (regular verbs)", "actual_content": "played/watched" },
  { "section": "Reading Comprehension", "learning_target": "making inferences", "actual_content": "because/so" },
  { "section": "Speaking", "learning_target": "giving opinions", "actual_content": "I like/I don't like" },

  {"section": "Vocabulary", "learning_target": "clothes", "actual_content": "shirt/pants, shoes/hat"},

  {"section": "Grammar", "learning_target": "perfect tense (regular verbs)", "actual_content": "have/had done/eaten"},
  {"section": "Reading Comprehension", "learning_target": "special announcements", "actual_content": "effective immediately..."},
]


print(len(esl_lesson_plans))


def insert_lesson_plans(filename, lesson_plans, start_row=9):
  """
  Inserts lesson plans into the original Excel document.

  Args:
    filename: The name of the original Excel file.
    lesson_plans: A list of dictionaries containing lesson plans.
    start_row: The starting row (inclusive) to insert data (default: 9).
  """
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    print(f"File not found: {filename}")
    return

  worksheet = workbook.active

  # Insert lesson plans starting from start_row
  row_index = start_row
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=4).value = "50 mins"
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  # Save the changes to the original document
  workbook.save(filename)
  print(f"Lesson plans inserted. Original file modified: {filename}")


# Example usage
filename = "[TLI]_Blizzard_Learning_program_25_classes.xlsx"

# Remember to replace this with your actual backup directory path
backup_directory = "/home/sweeneyphilip11/TLI_Schedule/backup_directory"

# Create a backup of the original file (consider using libraries like shutil for robust copying)
shutil.copy(filename, f"{backup_directory}/{filename}")  # Example using shutil.copy

insert_lesson_plans(filename, esl_lesson_plans)


