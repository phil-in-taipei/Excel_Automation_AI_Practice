from openpyxl import Workbook

def create_esl_schedule(lesson_plans):
  """
  Creates an Excel file (.xlsx) named "schedule.xlsx" with lesson plans.

  Args:
    lesson_plans: A list of dictionaries with "section", "learning_target",
                  and "actual_content" keys.
  """
  workbook = Workbook()
  worksheet = workbook.active

  # Set column headers
  worksheet.cell(row=1, column=3).value = "Section"
  worksheet.cell(row=1, column=5).value = "Learning Target"
  worksheet.cell(row=1, column=6).value = "Actual Content"

  # Start data from the second row
  row_index = 2
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  workbook.save("schedule.xlsx")

# Example usage
esl_lesson_plans = [
  # ... your list of dictionaries here ...
 {
    "section": "Vocabulary",
    "learning_target": "describing places",
    "actual_content": "big/small, clean/dirty"
  },
  {
    "section": "Grammar",
    "learning_target": "present continuous",
    "actual_content": "am/is/are doing"
  },
  {
    "section": "Reading Comprehension",
    "learning_target": "main idea",
    "actual_content": "who/what/where"
  },
  {
    "section": "Listening Comprehension",
    "learning_target": "numbers",
    "actual_content": "one/two/three"
  },
  {
    "section": "Speaking",
    "learning_target": "greetings",
    "actual_content": "Hello/Goodbye"
  },
  {
    "section": "Writing",
    "learning_target": "simple sentences",
    "actual_content": "Subject + verb"
  }
]

create_esl_schedule(esl_lesson_plans)

print("ESL lesson schedule created as 'schedule.xlsx'")
