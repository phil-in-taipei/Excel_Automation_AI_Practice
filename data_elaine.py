import openpyxl
import shutil

esl_customer_service_topics = [
   {
       "section": "Greeting Customers",
       "learning_target": "Welcoming customers with polite greetings",
       "actual_content": "Good morning! How can I assist you today?"
   },
   {
       "section": "Handling Complaints",
       "learning_target": "Responding to customer complaints professionally",
       "actual_content": "I apologize for the inconvenience. Please let me know how I can resolve this issue."
   },
   {
       "section": "Product Knowledge",
       "learning_target": "Providing detailed information about products or services",
       "actual_content": "This model comes with a one-year warranty and various color options."
   },
   {
       "section": "Upselling and Cross-selling",
       "learning_target": "Suggesting additional products or services to customers",
       "actual_content": "Would you be interested in adding this complementary accessory to your purchase?"
   },
   {
       "section": "Handling Refunds and Returns",
       "learning_target": "Explaining refund and return policies to customers",
       "actual_content": "You can return the item within 30 days with the original receipt for a full refund."
   },
   {
       "section": "Resolving Conflicts",
       "learning_target": "Defusing tense situations and finding solutions",
       "actual_content": "I understand your frustration. Let me escalate this to a manager for further assistance."
   },
   {
       "section": "Placing Orders",
       "learning_target": "Assisting customers with placing orders",
       "actual_content": "Please provide me with your shipping address and preferred payment method."
   },
   {
       "section": "Telephone Etiquette",
       "learning_target": "Using appropriate language and tone on the phone",
       "actual_content": "Thank you for your patience. How may I help you today?"
   },
   {
       "section": "Appointment Scheduling",
       "learning_target": "Scheduling appointments or reservations for customers",
       "actual_content": "I have availability on Tuesday at 2 PM or Thursday at 10 AM. Which works better for you?"
   },
   {
       "section": "Providing Directions",
       "learning_target": "Giving clear directions to customers",
       "actual_content": "Take a right at the next intersection, and our store will be on your left."
   },
   {
       "section": "Troubleshooting",
       "learning_target": "Guiding customers through troubleshooting steps",
       "actual_content": "Have you tried restarting the device? If that doesn't work, we can look into further solutions."
   },
   {
       "section": "Handling Payments",
       "learning_target": "Processing payments and explaining payment options",
       "actual_content": "We accept cash, credit cards, and mobile payments. How would you like to pay today?"
   },
   {
       "section": "Shipping and Delivery",
       "learning_target": "Explaining shipping and delivery processes",
       "actual_content": "Your order will be shipped within 2-3 business days via standard ground delivery."
   },
   {
       "section": "Loyalty Programs",
       "learning_target": "Informing customers about loyalty or reward programs",
       "actual_content": "As a valued customer, you can earn points for every purchase and redeem them for discounts."
   },
   {
       "section": "Closing Interactions",
       "learning_target": "Ending service interactions politely",
       "actual_content": "Thank you for your business. Please let me know if you need any further assistance."
   },
   {
       "section": "Follow-up and Feedback",
       "learning_target": "Requesting feedback and offering follow-up support",
       "actual_content": "Your feedback is valuable to us. Please fill out our short survey about your experience."
   },
   {
       "section": "Handling Difficult Customers",
       "learning_target": "Maintaining composure with challenging customers",
       "actual_content": "I understand this is frustrating, but let's try to resolve this calmly and professionally."
   },
   {
       "section": "Product Demonstrations",
       "learning_target": "Demonstrating product features and functionality",
       "actual_content": "Let me show you how this feature works and how it can benefit you."
   },
   {
       "section": "Sales Techniques",
       "learning_target": "Employing effective sales techniques",
       "actual_content": "This limited-time offer provides excellent value and meets your specific needs."
   },
   {
       "section": "Customer Retention",
       "learning_target": "Building customer loyalty and repeat business",
       "actual_content": "As a valued customer, you'll receive exclusive discounts and early access to new products."
   },
   {
       "section": "Managing Expectations",
       "learning_target": "Setting realistic expectations with customers",
       "actual_content": "Please allow 5-7 business days for delivery. I'll keep you updated on the status."
   },
   {
       "section": "Handling Emergencies",
       "learning_target": "Responding to emergency situations promptly",
       "actual_content": "For your safety, please follow the evacuation procedures immediately."
   },
   {
       "section": "Cross-cultural Communication",
       "learning_target": "Adapting communication styles across cultures",
       "actual_content": "In some cultures, maintaining eye contact is considered respectful during conversations."
   },
   {
       "section": "Resolving Misunderstandings",
       "learning_target": "Clarifying and resolving miscommunications",
       "actual_content": "Let me rephrase that to ensure we're on the same page."
   },
   {
       "section": "Suggesting Alternatives",
       "learning_target": "Offering alternative solutions or options",
       "actual_content": "If this product doesn't meet your needs, we have a similar model with additional features."
   }
]

print(len(esl_customer_service_topics))

def insert_lesson_plans(filename, lesson_plans, start_row=9):
  """
  Inserts lesson plans into the original Excel document.

  Args:
    filename: The name of the original Excel file.
    lesson_plans: A list of dictionaries containing lesson plans.
    start_row: The starting row (inclusive) to insert data (default: 9).
  """
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    print(f"File not found: {filename}")
    return

  worksheet = workbook.active

  # Insert lesson plans starting from start_row
  row_index = start_row
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=4).value = "50 mins"
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  # Save the changes to the original document
  workbook.save(filename)
  print(f"Lesson plans inserted. Original file modified: {filename}")


# Example usage
filename = "[TLI]_Blizzard_Learning_program_Lefay.xlsx"

# Remember to replace this with your actual backup directory path
#backup_directory = "/home/sweeneyphilip11/TLI_Schedule/backup_directory"
backup_directory = "/home/thinkpad/Documents/Python_Excel_Automation/Excel_Automation_AI_Practice/backup_directory"


# Create a backup of the original file (consider using libraries like shutil for robust copying)
shutil.copy(filename, f"{backup_directory}/{filename}")  # Example using shutil.copy

insert_lesson_plans(filename, esl_customer_service_topics)


def insert_lesson_plans(filename, lesson_plans, start_row=9):
  """
  Inserts lesson plans into the original Excel document.

  Args:
    filename: The name of the original Excel file.
    lesson_plans: A list of dictionaries containing lesson plans.
    start_row: The starting row (inclusive) to insert data (default: 9).
  """
  try:
    workbook = openpyxl.load_workbook(filename)
  except FileNotFoundError:
    print(f"File not found: {filename}")
    return

  worksheet = workbook.active

  # Insert lesson plans starting from start_row
  row_index = start_row
  for plan in lesson_plans:
    worksheet.cell(row=row_index, column=3).value = plan["section"]
    worksheet.cell(row=row_index, column=4).value = "50 mins"
    worksheet.cell(row=row_index, column=5).value = plan["learning_target"]
    worksheet.cell(row=row_index, column=6).value = plan["actual_content"]
    row_index += 1

  # Save the changes to the original document
  workbook.save(filename)
  print(f"Lesson plans inserted. Original file modified: {filename}")


# Example usage
filename = "[TLI]_Blizzard_Learning_program_Elaine.xlsx"

# Remember to replace this with your actual backup directory path
#backup_directory = "/home/sweeneyphilip11/TLI_Schedule/backup_directory"
backup_directory = "/home/thinkpad/Documents/Python_Excel_Automation/Excel_Automation_AI_Practice/backup_directory"


# Create a backup of the original file (consider using libraries like shutil for robust copying)
shutil.copy(filename, f"{backup_directory}/{filename}")  # Example using shutil.copy

insert_lesson_plans(filename, esl_customer_service_topics)
